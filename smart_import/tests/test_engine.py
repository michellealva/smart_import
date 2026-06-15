"""Tests for the Smart Import engine.

Uses core Frappe doctypes (User, Role, ToDo, DocType) only, so the suite runs
on any site without depending on CRM or other apps. Run with:

    bench --site <site> run-tests --module smart_import.tests.test_engine
"""

import unittest
from io import BytesIO

from openpyxl import load_workbook

from smart_import.engine import importer, mapper, reader, template


# --------------------------------------------------------------- reader (pure)
class TestReader(unittest.TestCase):
    def test_strip_sample_rows(self):
        sheet = {
            "name": "S",
            "headers": [reader.SAMPLE_COLUMN, "Email"],
            "rows": [["example", "a@x.com"], ["", "b@x.com"]],
        }
        out = reader._strip_sample_rows(sheet)
        self.assertEqual(out["headers"], ["Email"])  # marker column dropped
        self.assertEqual(out["rows"], [["b@x.com"]])  # example row dropped

    def test_strip_sample_rows_noop_without_marker(self):
        sheet = {"name": "S", "headers": ["Email"], "rows": [["a@x.com"]]}
        out = reader._strip_sample_rows(sheet)
        self.assertEqual(out["rows"], [["a@x.com"]])

    def test_noise_sheet_detection(self):
        self.assertTrue(reader._is_noise_sheet({"name": "Export Summary", "headers": ["x"], "rows": []}))
        self.assertTrue(reader._is_noise_sheet({"name": reader.LOOKUP_SHEET, "headers": ["x"], "rows": []}))
        self.assertTrue(
            reader._is_noise_sheet(
                {"name": "Sheet 1", "headers": ["This document was exported from Numbers."], "rows": []}
            )
        )
        self.assertFalse(reader._is_noise_sheet({"name": "User", "headers": ["Email"], "rows": []}))


# ---------------------------------------------------------- importer (pure)
class TestImporterPure(unittest.TestCase):
    def test_column_value_counts_case_insensitive(self):
        sheet = {"headers": ["Status"], "rows": [["Contract sent"], ["Contract Sent"], ["Won"]]}
        counts = importer.column_value_counts(sheet, {"Status": 0}, "Status")
        self.assertEqual(counts.get("Contract sent"), 2)  # case variants merged
        self.assertEqual(counts.get("Won"), 1)

    def test_coerce_numbers_and_check(self):
        self.assertEqual(importer.coerce({"fieldtype": "Int"}, "1,234"), 1234)
        self.assertEqual(importer.coerce({"fieldtype": "Currency"}, "$3.50"), 3.5)
        self.assertEqual(importer.coerce({"fieldtype": "Check"}, "yes"), 1)
        self.assertEqual(importer.coerce({"fieldtype": "Check"}, "no"), 0)
        self.assertEqual(importer.coerce({"fieldtype": "Data"}, "  hi  "), "hi")

    def test_coerce_select(self):
        col = {"fieldtype": "Select", "label": "P", "select_options": ["Low", "High"]}
        self.assertEqual(importer.coerce(col, "low"), "Low")
        with self.assertRaises(ValueError):
            importer.coerce(col, "Urgent")

    def test_resolve_select_cell(self):
        col = {"fieldtype": "Select", "label": "P", "select_options": ["Low", "High"]}
        self.assertEqual(importer.resolve_select_cell("low", col, {}), "Low")
        self.assertEqual(importer.resolve_select_cell("Urgent", col, {"urgent": "High"}), "High")
        self.assertIsNone(importer.resolve_select_cell("Urgent", col, {"urgent": "__blank__"}))
        with self.assertRaises(importer.SkipRow):
            importer.resolve_select_cell("Urgent", col, {"urgent": "__skip__"})

    def test_resolve_link_cell_match_blank_skip(self):
        col = {"link_doctype": "User"}
        # prepopulated map avoids any DB lookup
        self.assertEqual(
            importer.resolve_link_cell("admin@x", {"link_doctype": "User"}, {"User": {"admin@x": "Administrator"}}, {}, {}),
            "Administrator",
        )
        self.assertIsNone(
            importer.resolve_link_cell("ghost@x", col, {"User": {}}, {"ghost@x": "__blank__"}, {})
        )
        with self.assertRaises(importer.SkipRow):
            importer.resolve_link_cell("ghost@x", col, {"User": {}}, {"ghost@x": "__skip__"}, {})

    def test_resolve_link_cell_strict_no_fuzzy_no_autocreate(self):
        col = {"link_doctype": "User"}
        m = {"real@x": "Real User"}
        # exact (case-insensitive) match resolves
        self.assertEqual(
            importer.resolve_link_cell("real@x", col, {"User": dict(m)}, {}, {}), "Real User"
        )
        # a typed value that doesn't exist is NOT fuzzy-matched or created -> error
        with self.assertRaises(ValueError):
            importer.resolve_link_cell("Reallx", col, {"User": dict(m)}, {"reallx": "Reall User"}, {})
        # no decision + unknown value -> error (no silent auto-create)
        with self.assertRaises(ValueError):
            importer.resolve_link_cell("ghost@x", col, {"User": dict(m)}, {}, {})
        # an exact typed replacement resolves
        self.assertEqual(
            importer.resolve_link_cell("ghost@x", col, {"User": dict(m)}, {"ghost@x": "real@x"}, {}),
            "Real User",
        )

    def test_linkable_values_in_sheets(self):
        spec = {"entities": [{"id": "e0", "sheet": "User", "doctype": "User", "columns": []}]}
        data = {"sheets": [{"name": "User", "headers": ["Email"], "rows": [["loy@test.com"], ["x@y.com"]]}]}
        vals = importer.linkable_values_in_sheets(spec, data, "User")
        self.assertIn("loy@test.com", vals)
        self.assertNotIn("missing@z.com", vals)


# ------------------------------------------------------- mapper (needs site)
class TestMapper(unittest.TestCase):
    def test_doctype_from_sheet_name(self):
        self.assertEqual(mapper.doctype_from_sheet_name("User"), "User")
        self.assertEqual(mapper.doctype_from_sheet_name("Role"), "Role")
        self.assertIsNone(mapper.doctype_from_sheet_name("Definitely Not A Doctype 999"))

    def test_build_spec_maps_by_sheet_name(self):
        data = {"sheets": [{"name": "Role", "headers": ["Role Name"], "rows": [["Tester"]]}]}
        _, plan = mapper.build_spec(data)
        self.assertEqual(plan["entities"][0]["doctype"], "Role")

    def test_field_options_suggested(self):
        opts = {o["fieldname"]: o for o in mapper.field_options("Role")}
        self.assertIn("role_name", opts)
        self.assertTrue(opts["role_name"]["suggested"])

    def test_all_importable_doctypes(self):
        dts = mapper.all_importable_doctypes()
        self.assertIn("User", dts)
        self.assertIn("Role", dts)


# ----------------------------------------------------- validate (needs site)
class TestValidate(unittest.TestCase):
    def _spec(self, columns, doctype="ToDo", sheet="S"):
        return {"entities": [{"id": "e0", "sheet": sheet, "doctype": doctype, "columns": columns}]}

    def _issues(self, spec, data, kind):
        return [i for i in importer.validate(spec, data)["issues"] if i["type"] == kind]

    def test_flags_invalid_select(self):
        spec = self._spec([{"column": "P", "field": "priority", "fieldtype": "Select", "select_options": ["Low", "Medium", "High"]}])
        data = {"sheets": [{"name": "S", "headers": ["P"], "rows": [["Urgent"]]}]}
        sel = self._issues(spec, data, "select_values")
        self.assertTrue(sel)
        self.assertEqual([v["value"] for v in sel[0]["values"]], ["Urgent"])

    def test_select_value_dedup_case(self):
        spec = self._spec([{"column": "P", "field": "priority", "fieldtype": "Select", "select_options": ["Low"]}])
        data = {"sheets": [{"name": "S", "headers": ["P"], "rows": [["Urgent"], ["urgent"]]}]}
        sel = self._issues(spec, data, "select_values")[0]
        self.assertEqual(len(sel["values"]), 1)
        self.assertEqual(sel["values"][0]["count"], 2)

    def test_flags_missing_link(self):
        spec = self._spec([{"column": "U", "field": "x", "fieldtype": "Link", "link_doctype": "User"}])
        data = {"sheets": [{"name": "S", "headers": ["U"], "rows": [["smartimport-missing@example.invalid"]]}]}
        self.assertTrue(self._issues(spec, data, "link_values"))

    def test_link_ok_when_value_present_in_sibling_sheet(self):
        spec = {
            "entities": [
                {"id": "e0", "sheet": "Users", "doctype": "User", "columns": [{"column": "Email", "field": "email", "fieldtype": "Data"}]},
                {"id": "e1", "sheet": "T", "doctype": "ToDo", "columns": [{"column": "U", "field": "x", "fieldtype": "Link", "link_doctype": "User"}]},
            ]
        }
        val = "smartimport-newuser@example.invalid"
        data = {
            "sheets": [
                {"name": "Users", "headers": ["Email"], "rows": [[val]]},
                {"name": "T", "headers": ["U"], "rows": [[val]]},
            ]
        }
        # value is created by the User sheet -> not flagged
        self.assertFalse(self._issues(spec, data, "link_values"))

    def test_link_flagged_when_typo_not_in_sibling(self):
        spec = {
            "entities": [
                {"id": "e0", "sheet": "Users", "doctype": "User", "columns": [{"column": "Email", "field": "email", "fieldtype": "Data"}]},
                {"id": "e1", "sheet": "T", "doctype": "ToDo", "columns": [{"column": "U", "field": "x", "fieldtype": "Link", "link_doctype": "User"}]},
            ]
        }
        data = {
            "sheets": [
                {"name": "Users", "headers": ["Email"], "rows": [["smartimport-newuser@example.invalid"]]},
                {"name": "T", "headers": ["U"], "rows": [["a-totally-different-value@example.invalid"]]},
            ]
        }
        # referenced value exists neither in the site nor the sheet -> flagged
        self.assertTrue(self._issues(spec, data, "link_values"))

    def test_missing_required_value_flagged(self):
        # ToDo.description is mandatory with no default; a mapped-but-blank
        # column should be flagged as a missing_values issue.
        spec = self._spec([{"column": "Desc", "field": "description", "fieldtype": "Text Editor"}])
        data = {"sheets": [{"name": "S", "headers": ["Desc"], "rows": [[""], [""]]}]}
        mv = self._issues(spec, data, "missing_values")
        self.assertTrue(mv)
        self.assertEqual(mv[0]["count"], 2)

    def test_link_exact_match_not_flagged(self):
        spec = self._spec([{"column": "U", "field": "x", "fieldtype": "Link", "link_doctype": "User"}])
        data = {"sheets": [{"name": "S", "headers": ["U"], "rows": [["Administrator"]]}]}
        self.assertFalse(self._issues(spec, data, "link_values"))

    def test_link_near_miss_flagged_with_suggestion(self):
        # a near-miss of an existing record is flagged (not silently accepted),
        # and the close existing record is offered as the suggestion
        spec = self._spec([{"column": "U", "field": "x", "fieldtype": "Link", "link_doctype": "User"}])
        data = {"sheets": [{"name": "S", "headers": ["U"], "rows": [["Administratorr"]]}]}
        lv = self._issues(spec, data, "link_values")
        self.assertTrue(lv)
        sugg = {v["value"]: v["suggestion"] for v in lv[0]["values"]}
        self.assertEqual(sugg.get("Administratorr"), "Administrator")


# ------------------------------------------------- auto_creatable (needs site)
class TestAutoCreatable(unittest.TestCase):
    def test_user_not_creatable_role_is(self):
        # User needs email + first_name + real validation -> not creatable from a name
        self.assertFalse(importer.auto_creatable("User"))
        # Role is creatable from just its name
        self.assertTrue(importer.auto_creatable("Role"))


# ----------------------------------------------------- template (needs site)
class TestTemplate(unittest.TestCase):
    def test_discover_links_finds_doctype_link_and_flags_system(self):
        links = template.discover_links("ToDo")
        by_dt = {l["doctype"]: l for l in links}
        self.assertIn("DocType", by_dt)  # ToDo.reference_type -> DocType
        self.assertTrue(by_dt["DocType"]["system"])  # framework doctype
        self.assertIn("columns", by_dt["DocType"])

    def test_order_doctypes_dependency_first(self):
        order = template.order_doctypes(["ToDo", "DocType"])
        self.assertLess(order.index("DocType"), order.index("ToDo"))

    def test_build_workbook_blank(self):
        wb = load_workbook(BytesIO(template.build_workbook("Role", {}, [], "blank", {})))
        self.assertIn("Role", wb.sheetnames)
        self.assertGreaterEqual(wb["Role"].max_column, 1)

    def test_build_workbook_examples_has_marker_column(self):
        wb = load_workbook(BytesIO(template.build_workbook("Role", {}, [], "examples", {})))
        self.assertEqual(wb["Role"].cell(row=1, column=1).value, reader.SAMPLE_COLUMN)

    def test_build_workbook_select_gets_dropdown(self):
        # ToDo.priority is a Select field -> should get a data-validation dropdown
        wb = load_workbook(BytesIO(template.build_workbook("ToDo", {"ToDo": ["priority"]}, [], "blank", {})))
        self.assertTrue(list(wb["ToDo"].data_validations.dataValidation))

    def test_plan_shape(self):
        p = template.plan("Role")
        self.assertEqual(p["doctype"], "Role")
        self.assertIn("fields", p)
        self.assertIn("links", p)
