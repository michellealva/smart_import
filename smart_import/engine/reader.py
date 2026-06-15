"""Read .xlsx and .csv files into a simple structure:
{"sheets": [{"name": ..., "headers": [...], "rows": [[...], ...]}]}
"""

import csv

import frappe

# A column we add to "example rows" templates. Any row that still has a value
# in this column is treated as sample data and silently ignored on import, so
# users don't have to delete the examples we gave them.
SAMPLE_COLUMN = "⚠ Example rows — auto-skipped on import (add your data below)"
# earlier header text, still recognised so older downloaded templates import.
LEGACY_SAMPLE_COLUMNS = {"example rows — ignored on import (delete this column)"}

# Hidden sheet that holds dropdown lookup values in generated templates. It is
# never imported — skipped on read.
LOOKUP_SHEET = "__smart_import_lookups"


def get_path(file_url):
    file_doc = frappe.get_doc("File", {"file_url": file_url})
    return file_doc.get_full_path()


def _strip_sample_rows(sheet):
    """Drop example rows + the marker column from a sheet, in place-ish."""
    headers = sheet["headers"]
    marker = None
    for i, h in enumerate(headers):
        hl = str(h).strip().lower()
        if hl == SAMPLE_COLUMN.lower() or hl in LEGACY_SAMPLE_COLUMNS:
            marker = i
            break
    if marker is None:
        return sheet

    kept_rows = []
    for r in sheet["rows"]:
        flagged = marker < len(r) and r[marker] not in (None, "") and str(r[marker]).strip()
        if flagged:
            continue
        kept_rows.append([c for j, c in enumerate(r) if j != marker])

    sheet["headers"] = [h for j, h in enumerate(headers) if j != marker]
    sheet["rows"] = kept_rows
    return sheet


def _is_noise_sheet(sheet):
    """Sheets we should never try to import: our own lookup sheet, and the
    'Export Summary' sheet that Apple Numbers adds when exporting to .xlsx."""
    name = (sheet.get("name") or "").strip().lower()
    if name in (LOOKUP_SHEET.lower(), "export summary"):
        return True
    headers = sheet.get("headers") or []
    if headers and "exported from numbers" in str(headers[0]).lower():
        return True
    return False


def read_file(file_url):
    path = str(get_path(file_url))
    if path.lower().endswith(".csv"):
        sheets = [read_csv(path)]
    else:
        sheets = read_xlsx(path)
    sheets = [s for s in sheets if not _is_noise_sheet(s)]
    sheets = [_strip_sample_rows(s) for s in sheets]
    return {"sheets": [s for s in sheets if s["headers"] and s["rows"]]}


def read_csv(path):
    with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
        raw = list(csv.reader(f))

    raw = [r for r in raw if any(str(c).strip() for c in r)]
    headers = [str(h).strip() for h in (raw[0] if raw else [])]
    rows = raw[1:]

    keep = [i for i, h in enumerate(headers) if h]
    headers = [headers[i] for i in keep]
    rows = [[(r[i] if i < len(r) else None) for i in keep] for r in rows]

    return {"name": "Sheet1", "headers": headers, "rows": rows}


def read_xlsx(path):
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        headers = []
        rows = []
        for row in ws.iter_rows(values_only=True):
            if not headers:
                if row and any(c is not None and str(c).strip() for c in row):
                    headers = [str(c).strip() if c is not None else "" for c in row]
                continue
            if row and any(c is not None and str(c).strip() != "" for c in row):
                rows.append(list(row))

        keep = [i for i, h in enumerate(headers) if h]
        headers = [headers[i] for i in keep]
        rows = [[(r[i] if i < len(r) else None) for i in keep] for r in rows]

        sheets.append({"name": ws.title, "headers": headers, "rows": rows})
    wb.close()
    return sheets
