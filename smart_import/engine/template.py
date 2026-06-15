"""Build downloadable multi-sheet import templates for any doctype.

When you pick doctype A, any doctype it links to (B, and what B links to, ...)
can be included as its own sheet in the same workbook, ordered so linked records
come first. App-agnostic — works for any Frappe doctype.
"""

import datetime
from io import BytesIO

import frappe
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from smart_import.engine import reader

# value written in the marker column for every example row (any non-empty value
# makes the reader skip the row; the text is just a human hint)
SAMPLE_CELL = "example — delete row"
from smart_import.engine.mapper import (
    SKIP_FIELDTYPES,
    child_tables,
    field_options,
    importable_fields,
)

MAX_DISCOVERED = 40
INVALID_SHEET_CHARS = ":\\/?*[]"


def _is_system(doctype):
    """A doctype that belongs to the Frappe framework itself (vs. an app)."""
    module = frappe.db.get_value("DocType", doctype, "module")
    if not module:
        return True
    app = frappe.local.module_app.get(frappe.scrub(module))
    return app in (None, "frappe")


def _link_targets(doctype):
    for df in frappe.get_meta(doctype).fields:
        if df.fieldtype == "Link" and df.options:
            yield df


def _all_link_target_names(doctype):
    """Doctypes this one references — via its own Link fields AND the Link fields
    inside its child tables — so child-link masters get ordered first too."""
    seen = set()
    for df in frappe.get_meta(doctype).fields:
        if df.fieldtype == "Link" and df.options and df.options not in seen:
            seen.add(df.options)
            yield df.options
        elif df.fieldtype == "Table" and df.options:
            for cdf in frappe.get_meta(df.options).fields:
                if cdf.fieldtype == "Link" and cdf.options and cdf.options not in seen:
                    seen.add(cdf.options)
                    yield cdf.options


def _linkish_fields(doctype):
    """Link fields of a doctype AND the Link fields inside its child tables.

    Yields (docfield, via_table_label) — via_table_label is None for a direct
    link, or the child table's label when the link lives on a line item.
    """
    for df in frappe.get_meta(doctype).fields:
        if df.fieldtype == "Link" and df.options:
            yield df, None
        elif df.fieldtype == "Table" and df.options:
            for cdf in frappe.get_meta(df.options).fields:
                if cdf.fieldtype == "Link" and cdf.options:
                    yield cdf, (df.label or df.fieldname)


def discover_links(root):
    """Breadth-first walk over Link fields (direct and via child tables).

    Returns the linked doctypes (not root) in discovery order. System
    (framework) doctypes are reported but not expanded further, so the tree
    stays bounded and meaningful.
    """
    seen = {root}
    found = []
    queue = [root]
    while queue and len(found) < MAX_DISCOVERED:
        current = queue.pop(0)
        for df, via_table in _linkish_fields(current):
            target = df.options
            if target in seen or not frappe.db.exists("DocType", target):
                continue
            meta = frappe.get_meta(target)
            if meta.istable or meta.issingle:
                continue
            seen.add(target)
            system = _is_system(target)
            # a child-table link never *forces* the parent — it's opt-in
            mandatory = bool(df.reqd) and not via_table
            found.append(
                {
                    "doctype": target,
                    "via_field": df.fieldname,
                    "via_label": df.label or df.fieldname,
                    "from_doctype": current,
                    "system": system,
                    "mandatory": mandatory,
                    "via_table": bool(via_table),
                    # default column count for this tab (auto = identifier + required)
                    "columns": len(sheet_fields(target)),
                    # mandatory links always on; direct data links on; system and
                    # child-table (line-item) masters off by default (opt-in)
                    "default": mandatory or (not system and not via_table),
                }
            )
            if not system:
                queue.append(target)
    return found


def order_doctypes(doctypes):
    """Dependency order: a doctype's link targets come before it."""
    dset = set(doctypes)
    placed = []
    visiting = set()
    done = set()

    def visit(dt):
        if dt in done or dt in visiting:
            return
        visiting.add(dt)
        for target in _all_link_target_names(dt):
            if target in dset and target != dt:
                visit(target)
        visiting.discard(dt)
        done.add(dt)
        placed.append(dt)

    for dt in doctypes:
        visit(dt)
    return placed


def _title_field(doctype):
    meta = frappe.get_meta(doctype)
    autoname = meta.autoname or ""
    if autoname.startswith("field:"):
        return autoname.split(":", 1)[1]
    if meta.title_field and meta.title_field != "name":
        return meta.title_field
    return None


def sheet_fields(doctype, selected=None):
    """Columns for a sheet.

    Root sheet uses the user's selected fields (required always included).
    Linked sheets use an identifier (title) plus required fields, so the user
    can create the records that are referenced.
    """
    fields = importable_fields(doctype)
    if selected is not None:
        chosen = [df for df in fields if df.fieldname in selected or df.reqd]
        return chosen or fields

    out, seen = [], set()
    title = _title_field(doctype)
    for df in fields:
        if df.fieldname == title or df.reqd:
            if df.fieldname not in seen:
                out.append(df)
                seen.add(df.fieldname)
    if not out:
        # nothing identifying/required — fall back to the first few plain fields
        for df in fields:
            if df.fieldtype in ("Data", "Link", "Select"):
                out.append(df)
            if len(out) >= 4:
                break
    return out


def _fmt(value):
    if value is None:
        return ""
    if isinstance(value, (datetime.datetime, datetime.date)):
        return str(value)
    return value


def _sheet_name(doctype, used):
    name = doctype
    for ch in INVALID_SHEET_CHARS:
        name = name.replace(ch, " ")
    name = name.strip()[:31] or "Sheet"
    base, i = name, 1
    while name.lower() in used:
        suffix = " ({})".format(i)
        name = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(name.lower())
    return name


DROPDOWN_CAP = 2000  # don't build a dropdown for link targets bigger than this


class _LookupSheet:
    """Lazily-created hidden sheet that holds dropdown values too long/large to
    inline. Each call to add() takes a fresh column and returns a range formula."""

    def __init__(self, wb):
        self.wb = wb
        self.ws = None
        self.next_col = 1

    def add(self, values):
        if self.ws is None:
            self.ws = self.wb.create_sheet(title=reader.LOOKUP_SHEET)
            self.ws.sheet_state = "hidden"
        col = self.next_col
        self.next_col += 1
        for i, v in enumerate(values, start=1):
            self.ws.cell(row=i, column=col, value=v)
        letter = get_column_letter(col)
        return "'{0}'!${1}$1:${1}${2}".format(reader.LOOKUP_SHEET, letter, len(values))


def _apply_validations(ws, fields, offset, end_row, lookups):
    """Add Excel dropdowns for Select (choice) columns only.

    Link-field dropdowns are intentionally not generated for now — linked values
    are validated/created at upload time instead. Dropdowns are advisory
    (never reject typed values).
    """
    for i, df in enumerate(fields):
        if df.fieldtype != "Select":
            continue
        opts = [o.strip() for o in (df.options or "").split("\n") if o.strip()]
        if not opts:
            continue
        joined = ",".join(opts)
        if len(joined) <= 255 and not any("," in o for o in opts):
            formula = '"{}"'.format(joined)
        else:
            formula = lookups.add(opts)

        col = get_column_letter(i + 1 + offset)
        # showErrorMessage=False -> advisory: Excel suggests but allows any value
        dv = DataValidation(
            type="list", formula1=formula, allow_blank=True, showErrorMessage=False
        )
        ws.add_data_validation(dv)
        dv.add("{c}2:{c}{e}".format(c=col, e=end_row))


def _root_child_columns(root, child_map):
    """Selected line-item columns for the root sheet, flattened.

    Returns [(table_fieldname, child_doctype, docfield, header)]. Required and
    grid-visible (in_list_view) fields are force-included — the latter so the
    column shown in the line-item grid (e.g. the "Product" link) is always there,
    not just a secondary text field. Headers use the "Table: Field" convention
    the importer recognizes.
    """
    out = []
    targets = {ct["table_fieldname"]: ct for ct in child_tables(root)}
    autofill_text = {"Data", "Small Text", "Text", "Long Text"}
    for tf, fieldnames in (child_map or {}).items():
        ct = targets.get(tf)
        if not ct:
            continue
        chosen = set(fieldnames or [])
        if not chosen:
            continue
        fields = importable_fields(ct["child_doctype"])
        # if a grid Link identifies the line (e.g. Product), the import can fill a
        # required plain-text field (e.g. Product Name) from it — so don't clutter
        # the sheet with that redundant column unless the user explicitly chose it.
        has_list_link = any(f.fieldtype == "Link" and f.in_list_view for f in fields)
        for df in fields:
            # a required plain-text field the import auto-fills from the grid Link
            # is redundant in the sheet — always drop it (the Link column covers it)
            autofilled = (
                has_list_link
                and df.reqd
                and not df.in_list_view
                and df.fieldtype in autofill_text
            )
            if (df.fieldname in chosen or df.reqd or df.in_list_view) and not autofilled:
                header = "{}: {}".format(ct["label"], df.label or df.fieldname)
                out.append((tf, ct["child_doctype"], df, header))
    return out


def _link_placeholder(doctype, k=0):
    """A stable fake identifier for a linked record, e.g. "CRM Product 1".

    Used both as a Link cell's value and as the identifier of that linked
    record's own example sheet, so the example data lines up across sheets.
    """
    return "{} {}".format(doctype, k + 1)


def _real_link_value(doctype, k=0):
    """The k-th real record's display name for a link example, so a "Product"
    column shows actual product names (e.g. "Table") instead of "CRM Product 1".
    Returns None if there aren't enough real records to draw from."""
    title = _title_field(doctype) or "name"
    try:
        names = frappe.get_all(
            doctype, pluck=title, limit_page_length=k + 1, order_by="modified desc"
        )
    except Exception:
        return None
    return names[k] if k < len(names) else None


def _example_value(df, k=0):
    """A friendly placeholder for an example cell. `k` varies the value across
    rows/line items so they look distinct (and link consistently)."""
    ft = df.fieldtype
    if ft == "Link" and df.options:
        # prefer a real record name so the column reads naturally (e.g. a product
        # name), falling back to a placeholder when there are no records yet
        return _real_link_value(df.options, k) or _link_placeholder(df.options, k)
    if ft == "Int":
        return k + 1
    if ft in ("Float", "Currency", "Percent"):
        return (k + 1) * 100
    if ft == "Check":
        return 0
    if ft == "Date":
        return frappe.utils.today()
    if ft == "Select":
        opts = [o.strip() for o in (df.options or "").split("\n") if o.strip()]
        return opts[k % len(opts)] if opts else "example"
    return "example" if k == 0 else "example {}".format(k + 1)


def _placeholder_row(doctype, fields, k):
    """A full example row for a sheet that has no real records to show. The
    record's own identifier column gets the matching link placeholder, so other
    sheets that reference it (e.g. a Product line) point at a real-looking name."""
    title = _title_field(doctype)
    out = []
    for df in fields:
        if title and df.fieldname == title:
            out.append(_link_placeholder(doctype, k))
        else:
            out.append(_example_value(df, k))
    return out


def _write_example_rows(ws, root, parent_fields, child_cols):
    """Illustrative example rows for a flattened child-table template.

    Shows the FIRST parent repeated across three distinct line items (then a
    second parent with one), so "one line item per row, repeat the main columns,
    new parent = new record" reads at a glance. Line-item cells use real values
    from existing child records when available, otherwise distinct placeholders.
    """
    parent_fieldnames = [df.fieldname for df in parent_fields]
    primary_tf = child_cols[0][0]
    primary_dt = child_cols[0][1]
    primary_label = child_cols[0][3].split(":")[0].strip()
    primary_fields = [df.fieldname for (tf, cd, df, h) in child_cols if tf == primary_tf]

    # the column that ties a record's lines together (same as mapper.guess_group_key):
    # the parent's naming/title field if present, else the first parent column.
    title = _title_field(root)
    if title in parent_fieldnames:
        group_idx = parent_fieldnames.index(title)
    else:
        group_idx = 0 if parent_fieldnames else None

    parents = frappe.get_all(root, fields=["name"] + parent_fieldnames, limit_page_length=2) or [None]
    real_pool = (
        frappe.get_all(primary_dt, fields=primary_fields, limit_page_length=6)
        if primary_fields
        else []
    )

    def child_cells(k):
        sample = real_pool[k] if k < len(real_pool) else None
        cells = []
        for (tf, cd, df, header) in child_cols:
            if tf != primary_tf:
                cells.append("")
            elif sample and sample.get(df.fieldname) not in (None, ""):
                cells.append(_fmt(sample.get(df.fieldname)))
            else:
                cells.append(_example_value(df, k))
        return cells

    n = 0
    for i, p in enumerate(parents):
        pvals = []
        for df, fn in zip(parent_fields, parent_fieldnames):
            v = p.get(fn) if p else None
            # vary the placeholder per parent so two example records look distinct
            pvals.append(_fmt(v) if v not in (None, "") else _example_value(df, i))
        nlines = 3 if i == 0 else 1  # first parent shows the repeat
        for k in range(nlines):
            if k == 0:
                row_parent = list(pvals)
            else:
                # continuation line of the same record: keep ONLY the grouping
                # column (that repeat is what ties the lines together); blank the
                # rest so it reads as "same record, another line item".
                row_parent = ["" for _ in pvals]
                if group_idx is not None:
                    row_parent[group_idx] = pvals[group_idx]
            ws.append([SAMPLE_CELL] + row_parent + child_cells(k))
            n += 1

    # explain the layout via header comments (headers themselves must stay exact
    # so the importer can auto-map them)
    if group_idx is not None:
        gcell = ws.cell(row=1, column=2 + group_idx)  # +1 marker, +1 to 1-index
        gcell.comment = Comment(
            "Rows that share this value become ONE {} — repeat it on every line "
            "of the same record.".format(root),
            "Smart Import",
        )
    first_child_col = 2 + len(parent_fields)
    ws.cell(row=1, column=first_child_col).comment = Comment(
        "One {} line per row. Add more rows (repeating the grouping column) for "
        "more lines on the same record.".format(primary_label),
        "Smart Import",
    )
    return n


def _write_flattened_rows(ws, root, parent_fields, child_cols, row_filters, limit, mode):
    """Write one row per line item of the primary child table, repeating the
    parent values — the shape the flattened import expects. Other child tables'
    columns are left blank (single-table flattening is the common case)."""
    parent_fieldnames = [df.fieldname for df in parent_fields]
    primary_tf = child_cols[0][0]
    primary_dt = child_cols[0][1]
    primary_fields = [df.fieldname for (tf, cd, df, h) in child_cols if tf == primary_tf]

    recs = frappe.get_all(
        root, fields=["name"] + parent_fieldnames, filters=row_filters or {}, limit_page_length=limit
    )
    n = 0
    for rec in recs:
        pvals = [_fmt(rec.get(fn)) for fn in parent_fieldnames]
        lines = frappe.get_all(
            primary_dt,
            filters={"parent": rec.name, "parentfield": primary_tf, "parenttype": root},
            fields=["name"] + primary_fields,
            order_by="idx asc",
        )
        for line in lines or [None]:
            row = list(pvals)
            for (tf, cd, df, header) in child_cols:
                row.append(_fmt(line.get(df.fieldname)) if (line and tf == primary_tf) else "")
            if mode == "examples":
                row = ["example"] + row
            ws.append(row)
            n += 1
    return n


def build_workbook(root, fields_map, included, mode="blank", filters=None, child_map=None):
    """Return .xlsx bytes for the chosen doctype + its included linked sheets.

    fields_map: {doctype: [fieldnames]} chosen per sheet. A doctype absent from
    the map falls back to an auto-picked set (identifier + required fields).
    child_map: {table_fieldname: [fieldnames]} line-item columns to flatten into
    the root sheet (header "Table: Field").
    """
    filters = filters or {}
    fields_map = fields_map or {}
    child_map = child_map or {}
    included = [dt for dt in (included or []) if frappe.db.exists("DocType", dt)]
    doctypes = order_doctypes([root] + included)

    wb = Workbook()
    wb.remove(wb.active)
    used_names = set()
    lookups = _LookupSheet(wb)
    offset = 1 if mode == "examples" else 0

    # Pass 1 — write every sheet (so dropdowns can cross-reference each other).
    sheets = []
    for dt in doctypes:
        fields = sheet_fields(dt, fields_map.get(dt))
        fieldnames = [df.fieldname for df in fields]
        # line-item columns live on the root sheet only
        child_cols = _root_child_columns(dt, child_map) if dt == root else []
        ws = wb.create_sheet(title=_sheet_name(dt, used_names))

        headers = [df.label or df.fieldname for df in fields] + [c[3] for c in child_cols]
        # example rows get a marker column so the importer can ignore them
        if mode == "examples":
            headers = [reader.SAMPLE_COLUMN] + headers
        ws.append(headers)
        if mode == "examples":
            ws.cell(row=1, column=1).comment = Comment(
                "These example rows show the format and are skipped automatically "
                "on import.\n\nAdd your own rows below (leave this column blank), "
                "or just delete the example rows.",
                "Smart Import",
            )
        # parent + child fields in column order, for the dropdown pass
        validation_fields = fields + [c[2] for c in child_cols]

        datarows = 0
        if mode in ("records", "filtered", "examples"):
            # filters only make sense on the root doctype the user filtered
            row_filters = filters if (dt == root and mode == "filtered") else None
            # "examples" = just a few real rows to copy the format from
            limit = 5 if mode == "examples" else 0
            if child_cols and mode == "examples":
                datarows = _write_example_rows(ws, dt, fields, child_cols)
            elif child_cols:
                datarows = _write_flattened_rows(
                    ws, dt, fields, child_cols, row_filters, limit, mode
                )
            else:
                records = frappe.get_all(
                    dt,
                    fields=fieldnames or ["name"],
                    filters=row_filters or {},
                    limit_page_length=limit,
                )
                for rec in records:
                    row = [_fmt(rec.get(fn)) for fn in fieldnames]
                    if mode == "examples":
                        row = [SAMPLE_CELL] + row
                    ws.append(row)
                    datarows += 1
                # examples: if the system has few/no records, still show the shape
                # with placeholder rows (ignored on import via the marker column)
                if mode == "examples" and datarows < 3:
                    for k in range(datarows, 3):
                        ws.append([SAMPLE_CELL] + _placeholder_row(dt, fields, k))
                        datarows += 1

        sheets.append(
            {"dt": dt, "ws": ws, "fields": validation_fields, "datarows": datarows, "name": ws.title}
        )

    # Pass 2 — dropdowns for choice (Select) columns only.
    for s in sheets:
        _apply_validations(s["ws"], s["fields"], offset, max(1000, s["datarows"] + 200), lookups)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def plan(doctype):
    """Fields available for the root doctype + all discoverable linked sheets.

    `child_tables` are this doctype's line-item tables, offered as flattened
    columns in the root sheet (header "Table: Field").
    """
    return {
        "doctype": doctype,
        "fields": field_options(doctype),
        "links": discover_links(doctype),
        "child_tables": [
            {
                "table_fieldname": ct["table_fieldname"],
                "label": ct["label"],
                "child_doctype": ct["child_doctype"],
                "fields": field_options(ct["child_doctype"]),
            }
            for ct in child_tables(doctype)
        ],
    }
